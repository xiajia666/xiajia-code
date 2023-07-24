import openpyxl,json,os


class HandlerExcel():

    def get_excel_data_by_row(self, row,route):
        """带入一个行号，处理后返回这一行的数据是字典格式"""
        Appoint = openpyxl.load_workbook(route)
        example = Appoint.active
        dict_corresponding = {}

        for y in range(1, example.max_column + 1):
            corresponding = example.cell(row=1, column=y).value
            value = example.cell(row=row, column=y).value
            try:
                if corresponding == "verification_data":
                    pass
                else:
                    value = eval(value)
            except BaseException:
                pass
            dict_corresponding[corresponding] = value
        # if self.debug:
        #     print(("excel查询",dict_corresponding))
        #     self.info(("excel查询", dict_corresponding))
        return dict_corresponding

    def get_excel_data_by_id(self, excel_id, route=None):
        """定位id或定位行号，数字则时行号定位，非数字则是id定位"""
        if route == None:
            if "." not in excel_id:
                list_Route_id = excel_id.split("_", 1)
            else:
                list_Route_id = excel_id.split(".", 1)
                excel_id = list_Route_id[1]
            # self.info(("文件id拆分", list_Route_id))
            route = self.get_excel_path(list_Route_id[0])
        if isinstance(excel_id, int):
            return self.get_excel_data_by_row(excel_id,route)
        else:

            Appoint = openpyxl.load_workbook(route)
            example = Appoint.active

            for ids, i in enumerate(example.rows):
                if i[0].value == excel_id:
                    return self.get_excel_data_by_row(ids + 1, route)

    def get_excel_path(self, file_name):
        """匹配excel文件名称，返回地址"""
        def all_excel_route():
            data_cloud_route = os.path.abspath(os.path.abspath(os.path.join(os.path.dirname(__file__), "../"))) + "\\"+ r'data\cloud'
            list_all_excel_route = []
            for root, dirs, files in os.walk(data_cloud_route):
                for content in files:
                    if "__init__" not in content and content[-5:] == ".xlsx":
                        if content[:-5] == file_name:
                            list_all_excel_route.append(root + "\\" + content)

            if len(list_all_excel_route) != 1:
                return False
            return list_all_excel_route[0]

        return all_excel_route()

